from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PENDING_STATUSES = {"", "new", "pending", "open"}
DONE_STATUS = "done"
ERROR_STATUS = "error"


@dataclass
class TrackingCommand:
    row_number: int
    command_id: str
    command: str
    requested_by: str = ""


def load_pending_commands(config: dict[str, Any]) -> tuple[Path | None, list[TrackingCommand]]:
    if not config.get("enabled", True):
        return None, []
    path = _command_file(config)
    if path is None or not path.exists():
        return path, []

    sheet_name = str(config.get("sheet_name", "Commands"))
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers = _headers(ws)
    if "command" not in headers:
        return path, []

    commands: list[TrackingCommand] = []
    max_commands = int(config.get("max_commands_per_run", 5))
    for row_number in range(2, ws.max_row + 1):
        status = _cell(ws, row_number, headers, "status").strip().lower()
        command = _cell(ws, row_number, headers, "command").strip()
        if not command or status not in PENDING_STATUSES:
            continue
        command_id = _cell(ws, row_number, headers, "id").strip() or f"ROW-{row_number}"
        requested_by = _cell(ws, row_number, headers, "requestedby").strip()
        commands.append(TrackingCommand(row_number, command_id, command, requested_by))
        if len(commands) >= max_commands:
            break
    return path, commands


def update_command_result(config: dict[str, Any], command: TrackingCommand, status: str, response: str) -> None:
    path = _command_file(config)
    if path is None or not path.exists():
        return
    sheet_name = str(config.get("sheet_name", "Commands"))
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers = _headers(ws)
    _set_cell(ws, command.row_number, headers, "status", status)
    _set_cell(ws, command.row_number, headers, "processedat", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    _set_cell(ws, command.row_number, headers, "response", _clip(response, int(config.get("max_response_chars", 4000))))
    wb.save(path)


def _command_file(config: dict[str, Any]) -> Path | None:
    value = str(config.get("file", "") or "").strip()
    if not value:
        return None
    return Path(value).expanduser()


def _headers(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is None:
            continue
        headers[_normalize_header(value)] = col
    return headers


def _cell(ws, row: int, headers: dict[str, int], column: str) -> str:
    idx = headers.get(column)
    if not idx:
        return ""
    value = ws.cell(row, idx).value
    return "" if value is None else str(value)


def _set_cell(ws, row: int, headers: dict[str, int], column: str, value: str) -> None:
    idx = headers.get(column)
    if idx:
        ws.cell(row, idx).value = value


def _normalize_header(value: object) -> str:
    return str(value).strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def _clip(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."
