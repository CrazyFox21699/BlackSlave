from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
HEADER = "D9EAD3"
GROUP = "DDEBF7"
GOLD = "D9A300"
GREEN = "92D050"
YELLOW = "FFF200"
ORANGE = "F4B183"
WHITE = "FFFFFF"
GRAY = "BFBFBF"
BLACK = "000000"


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    sample_dir = root / "sample_data"
    sample_dir.mkdir(parents=True, exist_ok=True)
    path = sample_dir / "daily_tracking_master.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracking"
    ws.sheet_view.showGridLines = False

    _build_legend(ws)
    _build_headers(ws)
    _build_rows(ws)
    _style_sheet(ws)

    wb.save(path)
    print(path)


def _build_legend(ws) -> None:
    ws["R2"] = "Warning"
    ws["S2"] = "Delta >= 5%"
    ws["R3"] = "On-tracked"
    ws["S3"] = "Delta <5%"
    ws["R4"] = "Good-progress"
    ws["S4"] = "Delta >=0"
    for cell, color in [("R2", YELLOW), ("R3", BLUE), ("R4", GREEN)]:
        ws[cell].fill = PatternFill("solid", fgColor=color)
        ws[cell].font = Font(size=9, bold=True)
    for cell in ["S2", "S3", "S4"]:
        ws[cell].font = Font(size=9)


def _build_headers(ws) -> None:
    group_headers = {
        "G4:K4": "",
        "L4:L4": "PIC",
        "M4:M4": "Milestone",
        "N4:N4": "Start date\n(Plan)",
        "O4:O4": "End date\n(Plan)",
        "P4:P4": "Est (MH) OR\nSP",
        "Q4:Q4": "Target",
        "R4:R4": "Current\nProgress",
        "S4:S4": "Current\nProgress",
        "T4:T4": "Previous value",
        "U4:U4": "Delta",
        "V4:V4": "Note",
    }
    for rng, value in group_headers.items():
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = value
        cell.fill = PatternFill("solid", fgColor=GROUP if value else WHITE)
        cell.font = Font(size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Module", "1", "1", "0", "0", "",
        "Common", "Overall", "Level", "", "Item",
        "PIC", "Milestone", "Start date\n(Plan)", "End date\n(Plan)",
        "Est (MH) OR\nSP", "Target", "Current\nProgress", "Current\nProgress",
        "Previous value", "Delta", "Note", "Tu", "We",
    ]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=value)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(size=9, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["W6"] = 24
    ws["X6"] = 25
    for col in ["W", "X"]:
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=BLUE)
        ws[f"{col}5"].font = Font(color=WHITE, bold=True, size=9)
        ws[f"{col}6"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws[f"{col}6"].alignment = Alignment(horizontal="center")


def _build_rows(ws) -> None:
    rows = [
        ["MODULE", 1, 0, 0, 0, "", "Common", "Overall", "Level", "", "TASK 1", "", "", "2-Apr", "15-Jun", 213.2, "76%", "67%", "=P7*R7", "", "-10%", ""],
        ["MODULE", 1, 1, 0, 0, "", "Module", "1.1", 1, "", "MODULE 1", "", "PoM_M_1", "2-Apr", "15-Jun", 129.7, "74%", "70%", "=P8*R8", "", "-3%", ""],
        ["MODULE", 1, 1, 1, 0, "", "SPEC", "1.1.1", 2, "", "SPEC - action", "", "PoM_M_1", "2-Apr", "16-Apr", 65.9, "100%", "100%", "=P9*R9", "", "0%", ""],
        ["MODULE", 1, 1, 1, 1, "", "SPEC", "1.1.1.1", 3, "", "SPEC - action", "Lion", "PoM_M_1", "2-Apr", "6-Apr", 28.0, "100%", "100%", "=P10*R10", "", "0%", ""],
        ["MODULE", 1, 1, 1, 2, "", "SPEC", "1.1.1.2", 3, "", "SPEC - action", "Tiger", "PoM_M_1", "13-Apr", "14-Apr", 9.3, "100%", "100%", "=P11*R11", "", "0%", ""],
        ["MODULE", 1, 1, 1, 3, "", "SPEC", "1.1.1.3", 3, "", "SPEC - action", "Lion", "PoM_M_1", "13-Apr", "14-Apr", 5.6, "100%", "100%", "=P12*R12", "", "0%", ""],
        ["MODULE", 1, 1, 1, 4, "", "SPEC", "1.1.1.4", 3, "", "SPEC - action", "", "PoM_M_1", "16-Apr", "16-Apr", 9.3, "100%", "100%", "=P13*R13", "", "0%", ""],
        ["MODULE", 1, 1, 1, 5, "", "SPEC", "1.1.1.5", 3, "", "SPEC - action", "", "PoM_M_1", "16-Apr", "16-Apr", 5.6, "100%", "100%", "=P14*R14", "", "0%", ""],
        ["MODULE", 1, 1, 1, 6, "", "SPEC", "1.1.1.6", 3, "", "SPEC - action", "Cat", "PoM_M_1", "15-Apr", "15-Apr", 8.0, "100%", "100%", "=P15*R15", "", "0%", ""],
        ["MODULE", 1, 1, 2, 0, "", "SPEC", "1.1.2", 2, "", "SPEC - action", "", "PoM_M_1", "17-Apr", "21-Apr", 16.0, "100%", "100%", "=P16*R16", "", "0%", ""],
        ["MODULE", 1, 1, 2, 1, "", "SPEC", "1.1.2.1", 3, "", "SPEC - action", "", "PoM_M_1", "17-Apr", "21-Apr", 8.0, "100%", "100%", "=P17*R17", "", "0%", ""],
        ["MODULE", 1, 1, 2, 2, "", "SPEC", "1.1.2.2", 3, "", "SPEC - action", "", "PoM_M_1", "17-Apr", "21-Apr", 8.0, "100%", "100%", "=P18*R18", "", "0%", ""],
        ["MODULE", 1, 1, 3, 0, "", "SPEC-CUS", "1.1.3", 2, "", "SPEC-CUS - action", "", "PoM_M_4", "22-May", "26-May", 7.4, "57%", "0%", "=P19*R19", "", "-57%", "waiting OEM feedback"],
        ["MODULE", 1, 1, 3, 1, "", "SPEC-CUS", "1.1.3.1", 3, "", "SPEC-CUS - action", "Lion", "PoM_M_4", "22-May", "22-May", 4.2, "100%", "0%", "=P20*R20", "", "-100%", "blocked waiting input"],
        ["MODULE", 1, 1, 3, 2, "", "SPEC-CUS", "1.1.3.2", 3, "", "SPEC-CUS - action", "Tiger", "PoM_M_4", "25-May", "25-May", 0.6, "0%", "0%", "=P21*R21", "", "0%", ""],
        ["MODULE", 1, 1, 3, 3, "", "SPEC-CUS", "1.1.3.3", 3, "", "SPEC-CUS - action", "Lion", "PoM_M_4", "25-May", "25-May", 0.6, "0%", "0%", "=P22*R22", "", "0%", ""],
        ["MODULE", 1, 1, 3, 4, "", "SPEC-CUS", "1.1.3.4", 3, "", "SPEC-CUS - action", "", "PoM_M_4", "26-May", "26-May", 0.6, "0%", "0%", "=P23*R23", "", "0%", ""],
        ["MODULE", 1, 1, 3, 5, "", "SPEC-CUS", "1.1.3.5", 3, "", "SPEC-CUS - action", "", "PoM_M_4", "26-May", "26-May", 0.6, "0%", "0%", "=P24*R24", "", "0%", ""],
        ["MODULE", 1, 1, 3, 6, "", "SPEC-CUS", "1.1.3.6", 3, "", "SPEC-CUS - action", "Cat", "PoM_M_4", "26-May", "26-May", 0.6, "0%", "0%", "=P25*R25", "", "0%", ""],
        ["MODULE", 1, 1, 4, 0, "", "SPEC-CUS", "1.1.4", 2, "", "SPEC-CUS - action", "", "PoM_M_4", "27-May", "29-May", 1.3, "0%", "0%", "=P26*R26", "", "0%", ""],
        ["MODULE", 1, 1, 4, 1, "", "SPEC-CUS", "1.1.4.1", 3, "", "SPEC-CUS - action", "", "PoM_M_4", "27-May", "29-May", 0.6, "0%", "0%", "=P27*R27", "", "0%", ""],
        ["MODULE", 1, 1, 4, 2, "", "SPEC-CUS", "1.1.4.2", 3, "", "SPEC-CUS - action", "", "PoM_M_4", "27-May", "29-May", 0.6, "0%", "0%", "=P28*R28", "", "0%", ""],
        ["MODULE", 1, 1, 5, 0, "", "Code", "1.1.5", 2, "", "Code - action", "", "PoM_M_2", "14-Apr", "25-May", 20.5, "45%", "45%", "=P29*R29", "", "0%", ""],
        ["MODULE", 1, 1, 5, 1, "", "Code", "1.1.5.1", 3, "", "Code - action", "Cat", "PoM_M_2", "14-Apr", "16-Apr", 9.3, "100%", "100%", "=P30*R30", "", "0%", ""],
        ["MODULE", 1, 1, 5, 2, "", "Code", "1.1.5.2", 3, "", "Code - action", "Tiger", "PoM_M_2", "25-May", "25-May", 2.8, "0%", "0%", "=P31*R31", "", "0%", ""],
        ["MODULE", 1, 1, 5, 3, "", "Code", "1.1.5.3", 3, "", "Code - action", "Cat", "PoM_M_2", "25-May", "25-May", 2.8, "0%", "0%", "=P32*R32", "", "0%", ""],
        ["MODULE", 1, 1, 5, 4, "", "Code", "1.1.5.4", 3, "", "Code - action", "", "PoM_M_2", "25-May", "25-May", 2.8, "0%", "0%", "=P33*R33", 9.333333333, "0%", ""],
        ["MODULE", 1, 1, 5, 5, "", "Code", "1.1.5.5", 3, "", "Code - action", "Cat", "PoM_M_2", "25-May", "25-May", 2.8, "0%", "0%", "=P34*R34", "", "0%", ""],
        ["MODULE", 1, 1, 5, 6, "", "Code", "1.1.5.6", 3, "", "Code - action", "Cat", "PoM_M_2", "25-May", False, 2.8, "0%", "0%", "=P35*R35", 4.0, "0%", ""],
    ]
    start_row = 7
    for row_idx, values in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx in [17, 18, 21]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str) and cell.value.endswith("%"):
                cell.value = float(cell.value[:-1]) / 100

    ws["P6"] = "=SUM(P7:P35)"
    ws["R6"] = "=AVERAGE(R7:R35)"
    ws["S6"] = "=SUM(S7:S35)"
    ws["U6"] = "=AVERAGE(U7:U35)"
    for cell in ["P6", "R6", "S6", "U6"]:
        ws[cell].fill = PatternFill("solid", fgColor=GOLD)
        ws[cell].font = Font(bold=True)


def _style_sheet(ws) -> None:
    thin = Side(style="thin", color=BLACK)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = {
        "A": 11, "B": 4, "C": 4, "D": 4, "E": 4, "F": 4,
        "G": 12, "H": 10, "I": 7, "J": 4, "K": 18,
        "L": 10, "M": 12, "N": 11, "O": 11, "P": 12, "Q": 12,
        "R": 12, "S": 12, "T": 13, "U": 9, "V": 24, "W": 6, "X": 6,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.column_dimensions["J"].hidden = True
    ws.row_dimensions[4].height = 34
    ws.row_dimensions[5].height = 42
    ws.freeze_panes = "G7"
    ws.auto_filter.ref = "A5:X35"

    for row in ws.iter_rows(min_row=4, max_row=35, min_col=1, max_col=24):
        for cell in row:
            cell.border = border
            cell.font = Font(size=9, bold=cell.row in {4, 6})
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(7, 36):
        if row in {7, 8}:
            fill = PatternFill("solid", fgColor=GOLD if row == 7 else LIGHT_BLUE)
        elif row in {9, 16, 19, 26, 29}:
            fill = PatternFill("solid", fgColor="B8CCE4")
        else:
            fill = PatternFill("solid", fgColor=WHITE)
        for col in range(1, 25):
            ws.cell(row=row, column=col).fill = fill
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E2F0D9")
        ws.cell(row=row, column=18).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row=row, column=19).fill = PatternFill("solid", fgColor="DDEBF7")
        ws.cell(row=row, column=21).fill = PatternFill("solid", fgColor=GREEN)

    for row in range(7, 36):
        ws[f"R{row}"].number_format = "0%"
        ws[f"U{row}"].number_format = "0%"
        ws[f"S{row}"].number_format = "0.00"
        ws[f"P{row}"].number_format = "0.0"
    ws["R6"].number_format = "0%"
    ws["U6"].number_format = "0%"

    ws.conditional_formatting.add("U7:U35", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add("U7:U35", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add("R7:R35", CellIsRule(operator="lessThan", formula=["0.5"], fill=PatternFill("solid", fgColor=BLUE)))
    ws.conditional_formatting.add("R7:R35", CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", fgColor=BLUE)))


if __name__ == "__main__":
    main()
